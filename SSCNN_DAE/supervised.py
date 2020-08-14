# -*- coding: utf-8 -*-
"""
Created on Thu Aug 16 01:41:11 2018

@author: 32628
"""

import openpyxl
from skimage import io,transform
import glob
import os
import tensorflow as tf
import numpy as np
import time
import matplotlib.pyplot as plt
from sklearn.metrics import roc_curve, auc
#tf.reset_default_graph()
#path2="C:/Users/32628/.spyder-py3/COE/another/"
path = "C:/Users/56342/Desktop/val_data/noise/"
##savepath='D:/acc/model.ckpt'
def read_img(path):
    cate=[path+x for x in os.listdir(path) if os.path.isdir(path+x)]
    mat=[]
    labels=[]
    for idx,folder in enumerate(cate):
        for files in glob.glob(folder+'/*.npy'):
            f=np.load(files)
            mat.append(f)
            labels.append(idx)
    return np.asarray(mat,np.float32),np.asarray(labels,np.int32)
data,label=read_img(path)
num_example=data.shape[0]
arr=np.load('C:/Users/56342/Desktop/val_data/data_86.npy')
'''arr=np.arange(num_example)
np.random.shuffle(arr)
np.save('C:/Users/32628/.spyder-py3/COE/data.npy',arr)'''
data=data[arr]
label=label[arr]
#print(arr)
ratio=0.9
s=np.int(num_example*ratio)
x_train=data[:s]
y_train=label[:s]
x_val=data[s:]
y_val=label[s:]

enc=tf.placeholder(tf.float32,shape=[None,25,25,8],name='enc')
lab_=tf.placeholder(tf.int32,shape=[None,],name='lab_')

def inference(input_tensor, train, regularizer):
    nodes=25*25*8
    reshaped=tf.reshape(input_tensor,[-1,nodes])
    with tf.variable_scope('layer4-fc1'):
        fc1_weights = tf.get_variable("w", [nodes,512],
                                      initializer=tf.truncated_normal_initializer(stddev=0.1))
        if regularizer != None: tf.add_to_collection('losses', regularizer(fc1_weights))
        fc1_biases = tf.get_variable("b", [512], initializer=tf.constant_initializer(0.1))

        fc1 = tf.nn.relu(tf.matmul(reshaped, fc1_weights) + fc1_biases)
        if train: fc1 = tf.nn.dropout(fc1, 0.3)

##    with tf.variable_scope('layer5-fc2'):
##        fc2_weights = tf.get_variable("w", [1024, 512],
##                                      initializer=tf.truncated_normal_initializer(stddev=0.1))
##        if regularizer != None: tf.add_to_collection('losses', regularizer(fc2_weights))
##        fc2_biases = tf.get_variable("b", [512], initializer=tf.constant_initializer(0.1))
##
##        fc2 = tf.nn.relu(tf.matmul(fc1, fc2_weights) + fc2_biases)
##        if train: fc2 = tf.nn.dropout(fc2, 0.5)

    with tf.variable_scope('layer6-fc3'):
        fc3_weights = tf.get_variable("w", [512, 2],
                                      initializer=tf.truncated_normal_initializer(stddev=0.1))
        if regularizer != None: tf.add_to_collection('losses', regularizer(fc3_weights))
        fc3_biases = tf.get_variable("b", [2], initializer=tf.constant_initializer(0.1))
        logit = tf.matmul(fc1, fc3_weights) + fc3_biases
    return logit
#----------------------------------------网络结束---------------------------
regularizer = tf.contrib.layers.l2_regularizer(0.0001)
logits = inference(enc,False,regularizer)

#(小处理)将logits乘以1赋值给logits_eval，定义name，方便在后续调用模型时通过tensor名字调用输出tensor
b = tf.constant(value=1,dtype=tf.float32)
logits_eval = tf.multiply(logits,b,name='logits_eval') 

loss=tf.nn.sparse_softmax_cross_entropy_with_logits(logits=logits, labels=lab_)
train_op=tf.train.AdamOptimizer(learning_rate=0.0005).minimize(loss)
correct_prediction = tf.equal(tf.cast(tf.argmax(logits,1),tf.int32), lab_)    
acc= tf.reduce_mean(tf.cast(correct_prediction, tf.float32))

#定义一个函数，按批次取数据
def minibatches(inputs=None, targets=None, batch_size=None, shuffle=False):
    assert len(inputs) == len(targets)
    if shuffle:
        indices = np.arange(len(inputs))
        np.random.shuffle(indices)
    for start_idx in range(0, len(inputs) - batch_size + 1, batch_size):
        if shuffle:
            excerpt = indices[start_idx:start_idx + batch_size]
        else:
            excerpt = slice(start_idx, start_idx + batch_size)
        yield inputs[excerpt], targets[excerpt]

#训练和测试数据

n_epoch=100                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                         
batch_size=40
##saver=tf.train.Saver()
##sess=tf.Session()
##saver.restore(sess,'D:/disease/model.ckpt')
saver=tf.train.Saver()
sess=tf.Session()  
sess.run(tf.global_variables_initializer())
Acc_train=[]
Acc_vali=[]
train_Loss=[]
y_score=[]
for epoch in range(n_epoch):
    start_time = time.time()
    
    #training
    train_loss, train_acc, n_batch = 0, 0, 0
    for x_train_a, y_train_a in minibatches(x_train, y_train, batch_size, shuffle=True):
        _,err,ac=sess.run([train_op,loss,acc], feed_dict={enc: x_train_a, lab_: y_train_a})
        train_loss += err; train_acc += ac; n_batch += 1
    if epoch%1==0:
##        print("   train loss: %f" % (np.sum(train_loss)/ n_batch))
        print(epoch)
 
        
##        print("   train acc: %f" % (np.sum(train_acc)/ n_batch))
        
##        saver.save(sess,savepath)
##        saver = tf.train.import_meta_graph('D:/acc/model.ckpt.meta')
##        saver.restore(sess,tf.train.latest_checkpoint('D:/acc/'))
        graph = tf.get_default_graph()
        enc = graph.get_tensor_by_name("enc:0")
        feed_dict = {enc:x_val}
   
        logits = graph.get_tensor_by_name("logits_eval:0")
        
        classification_result = sess.run(logits,feed_dict)
        #print(classification_result)

        output = []
        output = tf.argmax(classification_result,1).eval(session=sess)
        #print(output)
        correct_prediction=tf.equal(y_val,output)
        accuracy=tf.reduce_mean(tf.cast(correct_prediction,tf.float32))
        print(classification_result.shape)
##        acc=sess.run(accuracy)
        print("   train loss: %f" % (np.sum(train_loss)/ n_batch))
        print("   train acc: %f" % (np.sum(train_acc)/ n_batch))    
        print('validation acc:',sess.run(accuracy))
    y_score=output
##    print(type(y_score))
##    print(type(y_val))
    #y_score.append(output)
    train_Loss.append(np.sum(train_loss)/ n_batch)  
    Acc_train.append(np.sum(train_acc)/ n_batch)
    Acc_vali.append(sess.run(accuracy))

f = open('C:/Users/56342/Desktop/val_data/myself.txt', 'w')
for j in range(len(y_val)):
    f.write(str(y_val[j])+' '+str(y_score[j])+'\n')
f.close()

wb = openpyxl.Workbook()
ws=wb['Sheet']
for i in range(len(train_Loss)):
    ws.cell(row=i+1,column=2).value=train_Loss[i]
    ws.cell(row=i+1,column=1).value=i
wb.save('C:/Users/56342/Desktop/val_data/loss_data_real4.xlsx')

wb1=openpyxl.Workbook()
ws1=wb['Sheet']
for i in range(len(train_Loss)):
    ws.cell(row=i+1,column=2).value=Acc_vali[i]
    ws.cell(row=i+1,column=1).value=i
wb.save('C:/Users/56342/Desktop/val_data/val_data_real4.xlsx')

fpr,tpr,threshold = roc_curve(y_val, y_score) ###计算真正率和假正率
'''wb1=openpyxl.Workbook()
ws1=wb['Sheet']
for i in range(len(fpr)):
    ws.cell(row=i+1,column=2).value=fpr[i]
    ws.cell(row=i+1,column=1).value=tpr[i]
wb.save('C:/Users/56342/Desktop/val_data/roc_data.xlsx')'''

roc_auc = auc(fpr,tpr)
print(roc_auc)
plt.figure(5)
plt.plot(fpr, tpr, color='darkorange',
         lw=2, label='ROC curve (area = %0.2f)' % roc_auc)
plt.xlabel('False Positive Rate')
plt.ylabel('True Positive Rate')
plt.title('Receiver operating characteristic example')
plt.legend(loc="lower right")

plt.figure(4)
x1_axis=range(n_epoch)
y1_axis=train_Loss
plt.plot(x1_axis,y1_axis, color='black', label='$DT$', linewidth=0.8)
plt.xticks(fontsize=20)
plt.yticks(fontsize=20)
plt.xlabel('Epoch',fontsize=20)
plt.ylabel('Train_loss',fontsize=20)
plt.title('Train_loss',fontsize=20)
plt.grid(True)
plt.figure(3)
plt.plot(x1_axis,Acc_train, color='red', label='$DT$', linewidth=2,marker='.')
plt.plot(x1_axis,Acc_vali, color='blue', label='$M_2$', linewidth=2,marker='.')
plt.xticks(fontsize=20)
plt.yticks(fontsize=20)
plt.xlabel('Epoch',fontsize=20)
plt.ylabel('Accuracy',fontsize=20)
plt.title('Accuracy',fontsize=20)
label=['Acc_train','Acc_vali']
plt.legend(label,loc='lower right',fontsize=20)
plt.grid(True)
plt.show()

sess.close()
    
